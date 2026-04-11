"""
从数据库中提取已识别文档的关键字段，写入归档文件目录 Excel
字段：档号、文号、责任者、题名、日期、页数、密级、备注
"""
import asyncio
import re
import json
from copy import copy
from pathlib import Path

import xlrd
import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine

# ===== 配置 =====
import os as _os
DATABASE_URL = _os.getenv("DATABASE_URL", "postgresql+asyncpg://postgres:123456@localhost:5432/ocr_db")
# 模版文件所在文件夹（用于按 file_path 模糊匹配数据库任务）
TEMPLATE_FOLDER = _os.getenv("TEMPLATE_FOLDER", r"D:\GOOLGE\软件著录\模版文件")
# 写入目标 Excel（xls 格式，使用 openpyxl 会生成新文件覆盖）
OUTPUT_XLSX = _os.getenv("OUTPUT_XLSX", r"D:\GOOLGE\软件著录\归档文件目录（所需字段）.xls")


# ===== 字段提取函数 =====
def _extract_archive_number(filename: str) -> str:
    stem = Path(filename).stem.strip()
    if not stem:
        return ""

    ws_match = re.match(r'^(WS[·.]?\d{4}[·.]?[A-Z]\d+(?:-\d+)+)$', stem, re.IGNORECASE)
    if ws_match:
        return ws_match.group(1)

    kj_match = re.match(r'^(KJ(?:-[A-Za-z0-9]+){4,})$', stem, re.IGNORECASE)
    if kj_match:
        return kj_match.group(1)

    legacy_ws_match = re.match(r'^(WS[·.]?\d{4}[·.]?[A-Z]\d+[-]\d+)$', stem, re.IGNORECASE)
    if legacy_ws_match:
        return legacy_ws_match.group(1)

    if re.match(r'^(KJ[-].*)$', stem, re.IGNORECASE):
        parts = stem.split('-')
        if len(parts) >= 5:
            return '-'.join(parts[:5])

    return ""


# 通用归档页表格标签 → 标准字段映射
# 归档页（公文处理单/发文处理单/收文登记等）可能出现在文件首页或末页
_TABLE_LABEL_MAP: dict[str, str] = {
    # 题名
    "文件标题": "题名", "标题": "题名", "题名": "题名", "事由": "题名",
    "文件名称": "题名", "公文标题": "题名",
    # 文号
    "原文号": "文号", "文号": "文号", "发文字号": "文号", "发文号": "文号",
    "来文文号": "文号", "来文字号": "文号",
    # 责任者
    "来文单位": "责任者", "发文单位": "责任者", "发文机关": "责任者",
    "主送单位": "责任者", "责任者": "责任者", "来文机关": "责任者",
    "印发单位": "责任者", "制发单位": "责任者",
    # 日期
    "收文日期": "日期", "成文日期": "日期", "发文日期": "日期",
    "印发日期": "日期", "签发日期": "日期",
    # 密级
    "密级": "密级", "秘密等级": "密级", "机密等级": "密级",
    # 备注
    "备注": "备注",
}


def _extract_from_table_data(result_json) -> dict[str, str]:
    """从 result_json 中的表格区域提取归档页结构化字段。

    扫描所有页面的 table 类型 region，根据 _TABLE_LABEL_MAP
    将表格单元格中的标签映射为标准归档字段。
    支持多列表格布局（如 [标签, 值, 标签, 值]）。
    """
    table_fields: dict[str, str] = {}
    if not result_json:
        return table_fields

    pages = result_json if isinstance(result_json, list) else [result_json]
    for page in pages:
        if not isinstance(page, dict):
            continue
        for region in page.get("regions", []):
            if region.get("type") != "table":
                continue
            table_data = region.get("table_data")
            if not isinstance(table_data, list):
                continue
            for row in table_data:
                if not isinstance(row, list) or len(row) < 2:
                    continue
                # 遍历行中的标签-值对（支持 [label, val, label, val, ...] 布局）
                i = 0
                while i < len(row) - 1:
                    cell_label = str(row[i]).strip()
                    cell_value = str(row[i + 1]).strip()
                    # 尝试精确匹配标签
                    target_field = _TABLE_LABEL_MAP.get(cell_label)
                    if target_field and cell_value and cell_value != "-":
                        # 优先保留第一个非空匹配（归档页通常只有一组值）
                        if target_field not in table_fields:
                            table_fields[target_field] = cell_value
                        i += 2
                    else:
                        i += 1
    return table_fields


def extract_fields(filename: str, full_text: str, result_json, page_count: int, *, file_path: str = "") -> dict:
    """从 OCR 结果中提取关键字段"""
    fields = {
        "档号": "",
        "文号": "",
        "责任者": "",
        "题名": "",
        "日期": "",
        "页数": str(page_count) if page_count else "",
        "密级": "",
        "备注": "",
    }

    if not full_text:
        full_text = ""

    # ===== 第一优先级：从结构化表格提取（归档页/公文处理单） =====
    table_fields = _extract_from_table_data(result_json)
    for field_name, value in table_fields.items():
        if value and field_name in fields:
            fields[field_name] = value

    # 所有文本（去除多余空格）
    text_clean = re.sub(r'\s+', ' ', full_text).strip()
    lines = [l.strip() for l in full_text.split('\n') if l.strip()]

    # --- 档号：从文件名提取 ---
    # 文件名格式如 WS·2024·D10-0311-001.jpg 或 KJ-JJ-2017-02-001-025.jpg
    if not fields["档号"]:
        fields["档号"] = _extract_archive_number(filename)

    # --- 文号：正则匹配常见公文文号格式（仅当表格未提取到时） ---
    if not fields["文号"]:
        wh_patterns = [
            r'[\u4e00-\u9fa5]+[\[〔\(（]?\d{4}[\]〕\)）]?\s*(?:第\s*)?\d+\s*号',
            r'[\u4e00-\u9fa5]{2,10}发[\[〔\(（]\d{4}[\]〕\)）]\d+号',
            r'[\u4e00-\u9fa5]{2,10}[\[〔\(（]\d{4}[\]〕\)）]\s*\d+\s*号',
        ]
        for pat in wh_patterns:
            m = re.search(pat, text_clean)
            if m:
                fields["文号"] = m.group(0).strip()
                break

    # --- 题名：提取文档标题（仅当表格未提取到时） ---
    if not fields["题名"] and result_json:
        pages = result_json if isinstance(result_json, list) else [result_json]
        for page in pages:
            if not isinstance(page, dict):
                continue
            for region in page.get("regions", []):
                rtype = region.get("type", "")
                content = region.get("content", "")
                if rtype in ("doc_title", "title", "paragraph_title", "content_title") and content:
                    if len(content) > len(fields["题名"]):
                        fields["题名"] = content.strip()

    # 如果没有从 regions 提取到标题，尝试从文本前几行找
    if not fields["题名"]:
        for line in lines[:10]:
            # 跳过太短或像页眉的行
            if len(line) >= 6 and not re.match(r'^第?\d+页', line):
                # 标题通常是"关于XXX的通知/决定/意见"等
                if re.search(r'(关于|通知|决定|意见|办法|规则|方法|规范|条例|规定)', line):
                    fields["题名"] = line
                    break

    # 还是没有，取最长的前几行
    if not fields["题名"] and lines:
        candidates = sorted(lines[:8], key=len, reverse=True)
        if candidates:
            fields["题名"] = candidates[0][:100]

    # --- 责任者：发文单位（仅当表格未提取到时） ---
    if not fields["责任者"]:
        resp_patterns = [
            # 页脚格式: "XX综合部  2024年6月18日印发" — 最可靠
            r'([\u4e00-\u9fa5]{4,30}(?:局|部|委员会|委|办|厅|院|会|中心|处|科|室))\s+\d{4}\s*年.*?印发',
            # 公司/集团+部门后缀 — 高可信度
            r'([\u4e00-\u9fa5]{4,30}(?:公司|集团)[\u4e00-\u9fa5]{0,6}(?:局|部|委|办|厅|院|会|中心|处|科|室))',
            # 通用实体+关于/发布/印发
            r'([\u4e00-\u9fa5]{2,20}(?:局|部|委员会|委|办|厅|院|会|中心|处|科|室))\s*(?:关于|发布|印发)',
            r'([\u4e00-\u9fa5]{4,20}(?:人民政府|人力资源|档案馆|档案局))',
        ]
        for pat in resp_patterns:
            m = re.search(pat, text_clean)
            if m:
                fields["责任者"] = m.group(1).strip()
                break

        # 从文号前缀推断责任者
        if not fields["责任者"] and fields["文号"]:
            m = re.match(r'([\u4e00-\u9fa5]+)', fields["文号"])
            if m:
                fields["责任者"] = m.group(1)

    # --- 日期（仅当表格未提取到时） ---
    if not fields["日期"]:
        date_patterns = [
            r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日',
            r'(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})',
            r'(\d{4})(\d{2})(\d{2})',
        ]
        for pat in date_patterns:
            m = re.search(pat, text_clean)
            if m:
                y, mo, d = m.group(1), m.group(2), m.group(3)
                fields["日期"] = f"{y}-{int(mo):02d}-{int(d):02d}"
                break

    # --- 密级（仅当表格未提取到时） ---
    if not fields["密级"]:
        mj_match = re.search(r'(绝密|机密|秘密|内部|公开)', text_clean[:200])
        if mj_match:
            fields["密级"] = mj_match.group(1)

    return fields


async def main():
    print("正在连接数据库...")
    engine = create_async_engine(DATABASE_URL)

    # 将 Windows 路径转为 SQL LIKE 模式（转义反斜线）
    folder_like = TEMPLATE_FOLDER.replace("\\", "\\\\") + "%"
    print(f"查询路径模式: {folder_like}")

    async with engine.begin() as conn:
        result = await conn.execute(text(
            "SELECT id, filename, mode, status, page_count, full_text, result_json, file_path "
            "FROM ocr_tasks "
            "WHERE file_path LIKE :pattern AND status = 'done' "
            "ORDER BY filename"
        ), {"pattern": folder_like})
        rows = result.fetchall()

        # 如果按路径查不到，则按最近任务查询（兜底）
        if not rows:
            print("按路径未匹配，尝试查询最近100条已完成任务...")
            result2 = await conn.execute(text(
                "SELECT id, filename, mode, status, page_count, full_text, result_json, file_path "
                "FROM ocr_tasks WHERE status = 'done' ORDER BY id DESC LIMIT 100"
            ))
            rows = result2.fetchall()

    await engine.dispose()

    print(f"共找到 {len(rows)} 条已识别任务（含重复）")
    if not rows:
        print("未找到任何已完成的任务，请先进行批量识别。")
        return

    # 按文件名去重，保留 id 最大（最新）的一条
    seen = {}
    for row in rows:
        task_id, filename, mode, status, page_count, full_text, result_json, file_path = row
        fname_key = Path(filename).name
        if fname_key not in seen or task_id > seen[fname_key][0]:
            seen[fname_key] = (task_id, filename, mode, status, page_count, full_text, result_json, file_path)

    deduped = sorted(seen.values(), key=lambda r: r[1])  # 按文件名排序
    print(f"去重后共 {len(deduped)} 条不同文件")

    # 提取字段
    headers = ['档号', '文号', '责任者', '题名', '日期', '页数', '密级', '备注']
    records = []
    for row in deduped:
        task_id, filename, mode, status, page_count, full_text, result_json, file_path = row
        fields = extract_fields(filename, full_text or "", result_json, page_count)
        records.append(fields)
        title_preview = fields['题名'][:30] if fields['题名'] else '（未提取到）'
        print(f"  [{task_id}] {filename}: 题名={title_preview}...")

    # 写入 Excel（openpyxl，保存为 .xls 同名文件但实际是 xlsx 格式）
    out_path = Path(OUTPUT_XLSX)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "归档文件目录"

    # 标题行
    ws.merge_cells('A1:H1')
    ws['A1'] = '归档文件目录'
    ws['A1'].font = openpyxl.styles.Font(size=14, bold=True)
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

    # 表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    # 数据行
    for row_idx, record in enumerate(records, 3):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))

    # 列宽
    col_widths = {'A': 20, 'B': 25, 'C': 15, 'D': 50, 'E': 12, 'F': 6, 'G': 8, 'H': 15}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # 保存（openpyxl 保存 .xls 扩展名，Excel 可直接打开）
    wb.save(str(out_path))
    print(f"\n已保存到: {out_path}")
    print(f"共写入 {len(records)} 条记录")
    print("\n注意：openpyxl 生成的是 xlsx 格式，文件扩展名为 .xls 但 Excel 可正常打开。")


if __name__ == "__main__":
    asyncio.run(main())
