"""
归档文件目录 Excel 导出服务
从 OCR 结果中提取关键字段，写入/追加到 Excel
"""
import re
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)

HEADERS = ['档号', '文号', '责任者', '题名', '日期', '页数', '密级', '备注']

DEFAULT_EXCEL_NAME = '归档文件目录.xlsx'

TITLE_TYPES = {'doc_title', 'title', 'paragraph_title', 'content_title', 'abstract_title', 'reference_title'}
TITLE_KEYWORDS = ('关于', '通知', '决定', '意见', '办法', '规则', '方法', '规范', '条例', '规定', '请示', '通报', '公告', '方案', '细则', '会议纪要')
ORG_SUFFIXES = (
    '工会委员会', '委员会办公室', '人力资源和社会保障局', '人力资源和社会保障厅', '人力资源和社会保障部',
    '人民政府', '总工会', '办公室', '工会', '委员会', '档案馆', '档案局', '有限责任公司', '集团有限公司',
    '股份有限公司', '有限公司', '检察院', '法院', '医院', '学校', '大学', '学院', '集团', '公司', '政府',
    '党委', '支部', '协会', '中心', '银行', '局', '厅', '部', '院', '馆'
)
ORG_SUFFIX_PATTERN = '|'.join(re.escape(item) for item in sorted(ORG_SUFFIXES, key=len, reverse=True))
ORG_BODY_PATTERN = rf'[\u4e00-\u9fa5A-Za-z0-9·（）()]{{2,60}}(?:{ORG_SUFFIX_PATTERN})'
RESP_HEAD_PATTERN = re.compile(rf'({ORG_BODY_PATTERN})\s*(?:关于|印发|发布|转发|公布|报送|请示|通知|决定|意见|办法|规定|通报|公告|方案)')
RESP_FULL_PATTERN = re.compile(rf'({ORG_BODY_PATTERN})$')
RESP_FRAGMENT_PATTERN = re.compile(rf'({ORG_BODY_PATTERN})')
DOC_NO_PATTERNS = (
    re.compile(r'([\u4e00-\u9fa5A-Za-z]{2,20}(?:字|发|函|办|通|报|党组|工)?(?:\[\d{4}\]|\(\d{4}\)|\d{4})\s*(?:第\s*)?\d+\s*号)'),
    re.compile(r'([\u4e00-\u9fa5A-Za-z]{2,20}(?:发|函|字|办)\s*(?:\[\d{4}\]|\(\d{4}\))\s*\d+\s*号)'),
)
DATE_PATTERN = re.compile(r'(\d{4})\s*(?:年|[./-])\s*(\d{1,2})\s*(?:月|[./-])\s*(\d{1,2})\s*日?')
CLASSIFICATION_PATTERN = re.compile(r'(绝密|机密|秘密|内部|公开|普通)')
PERIOD_DOC_NO_PATTERNS = (
    re.compile(r'(\d{4}\s*年\s*第\s*\d+\s*期)'),
    re.compile(r'(第\s*\d+\s*期)'),
)
ISSUED_BY_PATTERN = re.compile(r'([\u4e00-\u9fa5A-Za-z0-9·（）()]{4,40}(?:' + ORG_SUFFIX_PATTERN + r'))\s*\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日\s*(?:印发|发布|下发)')

def _extract_archive_number_from_path(file_path: str) -> str:
    """Parse folder hierarchy to extract archive number.
    Supports structures like: .../WS/2024/D30/0156/... → WS·2024·D30-0156
    """
    if not file_path:
        return ""
    parts = Path(file_path).parts
    for i in range(len(parts) - 1):
        upper = parts[i].upper()
        if upper in ('WS', 'KJ') and i + 3 < len(parts):
            category = upper
            year_part = parts[i + 1]
            retention_part = parts[i + 2]
            item_no_part = parts[i + 3]
            if re.fullmatch(r'\d{4}', year_part) and re.fullmatch(r'[A-Za-z]\d+[A-Za-z]?', retention_part):
                if re.fullmatch(r'\d+', item_no_part):
                    return f"{category}·{year_part}·{retention_part}-{item_no_part}"
    return ""

def _extract_archive_number(filename: str, file_path: str = "") -> str:
    stem = Path(filename).stem.strip()
    if not stem:
        return _extract_archive_number_from_path(file_path)

    normalized_stem = re.sub(r'\s+', '', stem)
    normalized_stem = normalized_stem.replace('•', '·').replace('・', '·')

    ws_match = re.match(r'^(WS[·.]?\d{4}[·.]?[A-Z]\d+-\d+)', normalized_stem, re.IGNORECASE)
    if ws_match:
        return ws_match.group(1)

    kj_match = re.match(r'^(KJ(?:-[A-Za-z0-9]+){4,})$', normalized_stem, re.IGNORECASE)
    if kj_match:
        return kj_match.group(1)

    legacy_ws_match = re.match(r'^(WS[·.]?\d{4}[·.]?[A-Z]\d+[·.]\d+)', normalized_stem, re.IGNORECASE)
    if legacy_ws_match:
        return legacy_ws_match.group(1)

    if re.match(r'^(KJ[-].*)$', normalized_stem, re.IGNORECASE):
        parts_list = normalized_stem.split('-')
        if len(parts_list) >= 5:
            return '-'.join(parts_list[:5])

    path_result = _extract_archive_number_from_path(file_path)
    if path_result:
        return path_result

    return ""

def _clean_line_text(text: str) -> str:
    clean = str(text or '').replace('\u3000', ' ').replace('\xa0', ' ')
    clean = re.sub(r'\s+', ' ', clean)
    return clean.strip(' \t\r\n，。；;:：')

def _normalize_search_text(text: str) -> str:
    clean = _clean_line_text(text)
    return (
        clean
        .replace('〔', '[')
        .replace('〕', ']')
        .replace('（', '(')
        .replace('）', ')')
        .replace('【', '[')
        .replace('】', ']')
    )

def _format_doc_no(text: str) -> str:
    clean = re.sub(r'\s+', '', _normalize_search_text(text))
    return clean.replace('[', '〔').replace(']', '〕').replace('(', '〔').replace(')', '〕')

def _bbox_to_rect(data: dict) -> list[float] | None:
    bbox = data.get('layout_bbox') or data.get('bbox') or []
    if not bbox:
        return None
    if isinstance(bbox[0], (list, tuple)):
        xs = [float(p[0]) for p in bbox if isinstance(p, (list, tuple)) and len(p) >= 2]
        ys = [float(p[1]) for p in bbox if isinstance(p, (list, tuple)) and len(p) >= 2]
        if xs and ys:
            return [min(xs), min(ys), max(xs), max(ys)]
        return None
    if len(bbox) >= 4:
        x1, y1, x2, y2 = [float(x) for x in bbox[:4]]
        return [min(x1, x2), min(y1, y2), max(x1, x2), max(y1, y2)]
    return None

def _get_page_dimensions(page: dict) -> tuple[float, float]:
    max_x = 0.0
    max_y = 0.0
    for region in page.get('regions', []):
        rect = _bbox_to_rect(region)
        if rect:
            max_x = max(max_x, rect[2])
            max_y = max(max_y, rect[3])
    for line in page.get('lines', []):
        rect = _bbox_to_rect(line)
        if rect:
            max_x = max(max_x, rect[2])
            max_y = max(max_y, rect[3])
    return max_x or 1.0, max_y or 1.0

def _build_page_items(page: dict, page_index: int, page_total: int) -> list[dict]:
    page_w, page_h = _get_page_dimensions(page)
    items = []

    for region in page.get('regions', []):
        if not isinstance(region, dict):
            continue
        text = _clean_line_text(region.get('content', ''))
        rtype = (region.get('type', '') or '').strip()
        if not text or rtype == 'table':
            continue
        rect = _bbox_to_rect(region) or [0.0, 0.0, page_w, page_h]
        x1, y1, x2, y2 = rect
        items.append({
            'text': text,
            'type': rtype or 'text',
            'source': 'region',
            'page_index': page_index,
            'page_total': page_total,
            'x1': x1,
            'y1': y1,
            'x2': x2,
            'y2': y2,
            'height': max(y2 - y1, 1.0),
            'y_ratio': y1 / page_h if page_h else 0.0,
        })

    for line in page.get('lines', []):
        if not isinstance(line, dict):
            continue
        text = _clean_line_text(line.get('text', ''))
        if not text:
            continue
        rect = _bbox_to_rect(line) or [0.0, 0.0, page_w, page_h]
        x1, y1, x2, y2 = rect
        items.append({
            'text': text,
            'type': 'line',
            'source': 'line',
            'page_index': page_index,
            'page_total': page_total,
            'x1': x1,
            'y1': y1,
            'x2': x2,
            'y2': y2,
            'height': max(y2 - y1, 1.0),
            'y_ratio': y1 / page_h if page_h else 0.0,
        })

    seen = set()
    deduped = []
    for item in sorted(items, key=lambda value: (value['y1'], value['x1'], 0 if value['source'] == 'region' else 1)):
        key = (
            item['page_index'],
            re.sub(r'\s+', '', item['text']),
            round(item['y1'] / 12),
            round(item['x1'] / 12),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped

def _build_text_only_items(full_text: str) -> list[dict]:
    lines = []
    for raw in full_text.split('\n'):
        clean = _clean_line_text(raw)
        if not clean or re.match(r'^---\s*第\s*\d+\s*页\s*---$', clean):
            continue
        lines.append(clean)
    page_h = float(max(len(lines), 1))
    items = []
    for idx, line in enumerate(lines):
        items.append({
            'text': line,
            'type': 'text',
            'source': 'text',
            'page_index': 0,
            'page_total': 1,
            'x1': 0.0,
            'y1': float(idx),
            'x2': 100.0,
            'y2': float(idx) + 1.0,
            'height': 1.0,
            'y_ratio': float(idx) / page_h if page_h else 0.0,
        })
    return items

def _collect_items(result_json, full_text: str) -> list[dict]:
    """
    将 OCR 的深层 JSON 数据结构（页、区域、行、字符）拍平并重组成一维数组。
    
    目标：标准化各种来源（PaddleOCR、Baidu VL等）输出的碎片坐标，方便基于 y 坐标的 `_score` 逻辑
    判断元素的真实物理位置分布。
    """
    if isinstance(result_json, list):
        pages = [page for page in result_json if isinstance(page, dict)]
    elif isinstance(result_json, dict):
        pages = [result_json]
    else:
        pages = []

    items = []
    for page_index, page in enumerate(pages):
        items.extend(_build_page_items(page, page_index, len(pages)))
    if not items:
        items = _build_text_only_items(full_text)
    return sorted(items, key=lambda value: (value['page_index'], value['y1'], value['x1']))

def _extract_doc_no_from_text(text: str) -> str:
    search_text = _normalize_search_text(text)
    for pattern in DOC_NO_PATTERNS:
        match = pattern.search(search_text)
        if match:
            return _format_doc_no(match.group(1))
    return ''

def _extract_date_candidates(text: str) -> list[str]:
    values = []
    search_text = _normalize_search_text(text)
    for match in DATE_PATTERN.finditer(search_text):
        year = int(match.group(1))
        month = int(match.group(2))
        day = int(match.group(3))
        if not (1900 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31):
            continue
        value = f'{year:04d}-{month:02d}-{day:02d}'
        if value not in values:
            values.append(value)
    return values

def _looks_like_page_number(text: str) -> bool:
    search_text = _normalize_search_text(text)
    return bool(re.fullmatch(r'(?:第?\s*\d+\s*页|共?\s*\d+\s*页|[-—－]+\s*\d+\s*[-—－]+|\d+\s*/\s*\d+)', search_text))

def _is_probable_title_text(text: str) -> bool:
    return any(keyword in text for keyword in TITLE_KEYWORDS)

def _looks_like_archive_stamp_code(text: str) -> bool:
    search_text = re.sub(r'\s+', '', _normalize_search_text(text)).upper()
    if not search_text:
        return False
    if re.search(r'全宗号|门类|年度|件号|保管期限|页数|目录号', text):
        return True
    if ("WS" in search_text or "KJ" in search_text) and sum(ch.isdigit() for ch in search_text) >= 6:
        return True
    if re.fullmatch(r'[A-Z0-9.·-]{10,}', search_text) and sum(ch.isdigit() for ch in search_text) >= 5:
        return True
    return False

def _is_generic_short_heading(text: str) -> bool:
    clean = re.sub(r'\s+', '', _clean_line_text(text))
    return clean in {"会议纪要", "通知", "决定", "意见", "办法", "通报", "公告", "方案", "细则"}

def _score_title_item(item: dict) -> int:
    """
    打分系统（标题评判）
    
    由于OCR识别出的标题可能因字体大小被割裂或与其他正文混淆，这里通过多个特征维度进行加减分：
    - 是否出现在首页顶部？（极高加分）
    - 是否包含类似页码、文号等干扰项？（一票否决/大幅扣分）
    - OCR是否将类型直接分类为 "title/doc_title"？（系统信任）
    
    返回分数越高，该行文本是文件标题的概率越大。
    """
    text = item['text']
    if item['page_index'] != 0 or item['y_ratio'] > 0.72:
        return -100
    if _looks_like_archive_stamp_code(text):
        return -100
    if re.search(r'全宗号|门类.*年度|件\s*号|保管期限|页\s*数', text):
        return -100
    if re.match(r'^[甲乙丙丁]方[:：]', text):
        return -100
    if re.match(r'^(第[一二三四五六七八九十百零\d]+条|[（(]?[一二三四五六七八九十百零\d]+[）).、])', text):
        return -100
    if _looks_like_page_number(text):
        return -100
    if _extract_doc_no_from_text(text):
        return -80
    if _extract_period_doc_no(text) and len(text) <= 24:
        return -60
    if _extract_date_candidates(text) and len(text) <= 24:
        return -40
    if CLASSIFICATION_PATTERN.fullmatch(text):
        return -40
    if _is_generic_short_heading(text):
        return -20

    score = 0
    if item['type'] in TITLE_TYPES:
        score += 14
    if item['source'] == 'region':
        score += 3
    if _is_probable_title_text(text):
        score += 10
    if 6 <= len(text) <= 40:
        score += 5
    elif len(text) <= 80:
        score += 2
    else:
        score -= 4
    if item['y_ratio'] < 0.35:
        score += 4
    elif item['y_ratio'] <= 0.68:
        score += 2
    if _is_probable_title_text(text) and any(suffix in text for suffix in ORG_SUFFIXES):
        score += 10
    if any(text.endswith(suffix) for suffix in ORG_SUFFIXES) and '关于' not in text:
        score -= 5
    if sum(ch.isdigit() for ch in text) >= 8:
        score -= 4
    return score

def _join_title_group(group: list[dict]) -> str:
    text = ''.join(item['text'].strip() for item in group if item['text'].strip())
    return re.sub(r'\s+', '', text)[:120]

def _extract_title(items: list[dict], fallback_lines: list[str]) -> str:
    top_items = [item for item in items if item['page_index'] == 0 and item['y_ratio'] <= 0.72]
    candidates = [item for item in top_items if _score_title_item(item) > 0]
    if candidates:
        groups = []
        current = []
        for item in sorted(candidates, key=lambda value: (value['y1'], value['x1'])):
            if not current:
                current = [item]
                continue
            prev = current[-1]
            gap = item['y1'] - prev['y2']
            if gap <= max(18.0, prev['height'] * 1.8):
                current.append(item)
            else:
                groups.append(current)
                current = [item]
        if current:
            groups.append(current)

        best_text = ''
        best_score = -10**9
        for group in groups[:6]:
            group_text = _join_title_group(group)
            if not group_text:
                continue
            score = sum(_score_title_item(item) for item in group)
            if _looks_like_archive_stamp_code(group_text):
                score -= 120
            if _is_probable_title_text(group_text):
                score += 6
            if len(group) >= 2:
                score += 6
            if 8 <= len(group_text) <= 80:
                score += 4
            if _is_probable_title_text(group_text) and any(suffix in group_text for suffix in ORG_SUFFIXES):
                score += 18
            if _is_generic_short_heading(group_text):
                score -= 16
            if group_text.endswith(('通知', '决定', '意见', '办法', '规则', '方法', '规范', '条例', '规定', '请示', '通报', '公告', '方案', '细则', '会议纪要')):
                score += 2
            if score > best_score:
                best_score = score
                best_text = group_text
        for index in range(min(len(groups) - 1, 5)):
            combined_text = _join_title_group(groups[index] + groups[index + 1])
            if not combined_text:
                continue
            if _looks_like_archive_stamp_code(combined_text):
                continue
            score = sum(_score_title_item(item) for item in groups[index] + groups[index + 1])
            if _is_probable_title_text(combined_text):
                score += 10
            if any(suffix in combined_text for suffix in ORG_SUFFIXES):
                score += 10
            if _is_probable_title_text(combined_text) and any(suffix in combined_text for suffix in ORG_SUFFIXES):
                score += 20
            if 12 <= len(combined_text) <= 100:
                score += 8
            if score > best_score:
                best_score = score
                best_text = combined_text
        if best_text:
            return best_text

    for line in fallback_lines[:10]:
        clean = _clean_line_text(line)
        if (
            len(clean) >= 6
            and not _looks_like_page_number(clean)
            and not _extract_doc_no_from_text(clean)
            and not _looks_like_archive_stamp_code(clean)
            and not _is_generic_short_heading(clean)
            and _is_probable_title_text(clean)
        ):
            return clean[:120]

    candidates = sorted(
        (
            _clean_line_text(line)
            for line in fallback_lines[:8]
            if _clean_line_text(line)
            and not _looks_like_archive_stamp_code(_clean_line_text(line))
            and not _looks_like_page_number(_clean_line_text(line))
            and not _extract_doc_no_from_text(_clean_line_text(line))
        ),
        key=len,
        reverse=True,
    )
    return candidates[0][:120] if candidates else ''

def _clean_org_name(text: str) -> str:
    clean = _clean_line_text(text)
    clean = re.sub(r'[（(][^()（）]*(?:盖章|印章|公章|章)[^()（）]*[）)]', '', clean)
    clean = re.sub(r'(关于|印发|发布|转发|公布|报送|请示|通知|决定|意见|办法|规定|通报|公告|方案).*$','', clean)
    clean = clean.strip(' ，。；;:：')
    match = RESP_FRAGMENT_PATTERN.search(clean)
    if match:
        clean = match.group(1).strip(' ，。；;:：')
    if 4 <= len(clean) <= 40:
        return clean
    return ''

def _extract_responsible_candidates(text: str) -> list[tuple[str, str]]:
    """Return list of (candidate, source_type) tuples.
    source_type: 'party' | 'head' | 'full' | 'issued' | 'fragment'
    """
    clean = _clean_line_text(text)
    candidates = []
    seen = set()
    party_match = re.match(r'^[甲乙丙丁]方[:：]\s*([\u4e00-\u9fa5A-Za-z0-9·（）()]+(?:有限公司|集团有限公司|集团|公司|委员会|办公室|档案局|档案馆|政府|中心|银行|局|厅|部|院|馆))', clean)
    if party_match:
        candidate = _clean_org_name(party_match.group(1))
        if candidate and candidate not in seen:
            candidates.append((candidate, 'party'))
            seen.add(candidate)
    issued_match = ISSUED_BY_PATTERN.search(clean)
    if issued_match:
        candidate = _clean_org_name(issued_match.group(1))
        if candidate and candidate not in seen:
            candidates.append((candidate, 'issued'))
            seen.add(candidate)
    for match in RESP_HEAD_PATTERN.finditer(clean):
        candidate = _clean_org_name(match.group(1))
        if candidate and candidate not in seen:
            candidates.append((candidate, 'head'))
            seen.add(candidate)
    for match in RESP_FULL_PATTERN.finditer(clean):
        candidate = _clean_org_name(match.group(1))
        if candidate and candidate not in seen:
            candidates.append((candidate, 'full'))
            seen.add(candidate)
    for match in RESP_FRAGMENT_PATTERN.finditer(clean):
        candidate = _clean_org_name(match.group(1))
        if candidate and candidate not in seen:
            candidates.append((candidate, 'fragment'))
            seen.add(candidate)
    return candidates

def _has_adjacent_issued_footer(item: dict, items: list[dict]) -> bool:
    """Check if there's a nearby footer/line item on the same page containing '年...印发'.

    Handles the common case where OCR splits the org name and '印发' date
    into separate items, e.g.:
      footer y=0.92: "重庆XX公司综合部"
      footer y=0.93: "2024年6月18日印发"
    """
    for other in items:
        if other is item:
            continue
        if other['page_index'] != item['page_index']:
            continue
        if abs(other['y_ratio'] - item['y_ratio']) > 0.06:
            continue
        if re.search(r'\d{4}\s*年.*?(?:印发|发布|下发)', other['text']):
            return True
    return False


def _extract_responsible(items: list[dict], doc_no: str) -> str:
    best_value = ''
    best_score = -10**9
    for item in items:
        for candidate, source_type in _extract_responsible_candidates(item['text']):
            score = 0
            # ── Strong prerequisite evidence (甲方/乙方, 盖章, 印发) ──
            if source_type == 'party':
                score += 25
            if source_type == 'issued':
                score += 20
            if item['type'] == 'seal':
                score += 18
            if any(word in item['text'] for word in ('盖章', '印章', '公章')):
                score += 15
            if any(word in item['text'] for word in ('发文单位', '发文机关', '主送')):
                score += 15
            if any(word in item['text'] for word in ('印发', '发布', '下发')):
                score += 10
            # ── Adjacent footer "印发" detection (split footer) ──
            if item['type'] in ('footer', 'line') and item['y_ratio'] > 0.8 and _has_adjacent_issued_footer(item, items):
                score += 22
            # ── Position signals ──
            if item['page_index'] == item['page_total'] - 1 and item['y_ratio'] > 0.55:
                score += 12
            if item['page_index'] == 0 and item['y_ratio'] < 0.35:
                score += 4
            # ── Pattern quality ──
            if source_type == 'head':
                score += 6
            if source_type == 'full':
                score += 4
            if source_type == 'fragment':
                score -= 5
            # ── Length heuristics ──
            if 4 <= len(candidate) <= 24:
                score += 3
            if len(candidate) > 32:
                score -= 2
            # ── Negative signals: title-position text is likely not the responsible party ──
            if item['page_index'] == 0 and 0.15 < item['y_ratio'] < 0.45 and item.get('type') in TITLE_TYPES:
                score -= 12
            if any(kw in item['text'] for kw in TITLE_KEYWORDS) and source_type == 'fragment':
                score -= 10
            if _extract_doc_no_from_text(item['text']):
                score -= 3
            if any(word in candidate for word in ('附件', '目录', '日期', '编号')):
                score -= 8
            if score > best_score:
                best_score = score
                best_value = candidate
    # Require minimum confidence: fragment-only low-score matches are unreliable
    if best_value and best_score >= 3:
        return best_value
    if doc_no:
        match = re.match(r'([\u4e00-\u9fa5A-Za-z]{2,20})', doc_no)
        if match:
            return match.group(1)
    return ''

def _extract_period_doc_no(text: str) -> str:
    """Extract period/issue style document numbers like '2024年第13期'."""
    search_text = _normalize_search_text(text)
    for pattern in PERIOD_DOC_NO_PATTERNS:
        match = pattern.search(search_text)
        if match:
            return re.sub(r'\s+', '', match.group(1))
    return ''

def _extract_doc_no(items: list[dict], fallback_lines: list[str]) -> str:
    best_value = ''
    best_score = -10**9
    for item in items:
        candidate = _extract_doc_no_from_text(item['text'])
        if not candidate:
            candidate = _extract_period_doc_no(item['text'])
        if not candidate:
            continue
        score = 0
        if item['page_index'] == 0:
            score += 8
        if item['y_ratio'] < 0.35:
            score += 10
        elif item['y_ratio'] < 0.5:
            score += 4
        else:
            score -= 4
        if len(item['text']) <= 40:
            score += 2
        if item['type'] in TITLE_TYPES:
            score += 1
        if score > best_score:
            best_score = score
            best_value = candidate
    if best_value:
        return best_value
    for line in fallback_lines[:12]:
        candidate = _extract_doc_no_from_text(line)
        if not candidate:
            candidate = _extract_period_doc_no(line)
        if candidate:
            return candidate
    return ''

def _extract_date(items: list[dict], fallback_lines: list[str]) -> str:
    best_value = ''
    best_marker = None
    for item in items:
        candidates = _extract_date_candidates(item['text'])
        if not candidates:
            continue
        for candidate in candidates:
            score = 0
            if item['page_index'] == item['page_total'] - 1:
                score += 4
            if item['y_ratio'] > 0.6:
                score += 10
            elif item['page_index'] == 0 and item['y_ratio'] < 0.35:
                score += 4
            if len(item['text']) <= 24:
                score += 2
            if any(word in item['text'] for word in ('印发', '发布', '下发')):
                score += 8
            if any(word in item['text'] for word in ('成文', '日期')):
                score += 4
            if item['page_index'] == item['page_total'] - 1 and ISSUED_BY_PATTERN.search(item['text']):
                score += 10
            if any(word in item['text'] for word in ('起', '截至', '活动', '培训', '实施')):
                score -= 4
            if any(word in item['text'] for word in ('会议',)) and '印发' not in item['text']:
                score -= 4
            marker = (score, item['page_index'], item['y1'])
            if best_marker is None or marker > best_marker:
                best_marker = marker
                best_value = candidate
    if best_value:
        return best_value
    for line in fallback_lines:
        candidates = _extract_date_candidates(line)
        if candidates:
            return candidates[0]
    return ''

def _extract_classification(items: list[dict], full_text: str) -> str:
    best_value = ''
    best_score = -10**9
    for item in items:
        clean = _clean_line_text(item['text'])
        match = CLASSIFICATION_PATTERN.search(clean)
        if not match:
            continue
        value = match.group(1)
        if value == '普通' and '密级' not in clean and len(clean) > 12:
            continue
        score = 0
        if item['page_index'] == 0:
            score += 4
        if item['y_ratio'] < 0.2 or item['y_ratio'] > 0.75:
            score += 6
        if len(clean) <= 12:
            score += 4
        if '密级' in clean:
            score += 8
        if score > best_score:
            best_score = score
            best_value = value
    if best_value:
        return best_value
    classified_match = re.search(r'密级[：:\s]*(' + CLASSIFICATION_PATTERN.pattern[1:-1] + r')', _normalize_search_text(full_text)[:800])
    if classified_match:
        return classified_match.group(1)
    match = CLASSIFICATION_PATTERN.search(_normalize_search_text(full_text)[:600])
    return match.group(1) if match else ''

# 通用归档页表格标签 → 标准字段映射
# 归档页（公文处理单/发文处理单/收文登记等）可能出现在文件首页或末页
_TABLE_LABEL_MAP: dict[str, str] = {
    # 题名
    "文件标题": "题名", "标题": "题名", "题名": "题名", "事由": "题名",
    "文件名称": "题名", "公文标题": "题名",
    # 文号
    "原文号": "文号", "文号": "文号", "发文字号": "文号", "发文号": "文号",
    "来文文号": "文号", "来文字号": "文号",
    # 责任者
    "来文单位": "责任者", "发文单位": "责任者", "发文机关": "责任者",
    "主送单位": "责任者", "责任者": "责任者", "来文机关": "责任者",
    "印发单位": "责任者", "制发单位": "责任者",
    # 日期
    "收文日期": "日期", "成文日期": "日期", "发文日期": "日期",
    "印发日期": "日期", "签发日期": "日期",
    # 密级
    "密级": "密级", "秘密等级": "密级", "机密等级": "密级",
    # 备注
    "备注": "备注",
}

# 用于从合并单元格（标签与值连写）中提取值的前缀列表
_TABLE_LABEL_PREFIXES = sorted(_TABLE_LABEL_MAP.keys(), key=len, reverse=True)


def _extract_from_table_data(result_json) -> dict[str, str]:
    """从 result_json 中的表格区域提取归档页结构化字段（fallback 用）。

    扫描所有页面的 table 类型 region，根据 _TABLE_LABEL_MAP
    将表格单元格中的标签映射为标准归档字段。
    支持多列表格布局（如 [标签, 值, 标签, 值]）和标签值合并单元格。
    """
    table_fields: dict[str, str] = {}
    if not result_json:
        return table_fields

    pages = result_json if isinstance(result_json, list) else [result_json]
    for page in pages:
        if not isinstance(page, dict):
            continue
        for region in page.get("regions", []):
            if region.get("type") != "table":
                continue
            table_data = region.get("table_data")
            if not isinstance(table_data, list):
                continue
            for row in table_data:
                if not isinstance(row, list) or len(row) < 2:
                    continue
                # 遍历行中的标签-值对（支持 [label, val, label, val, ...] 布局）
                i = 0
                while i < len(row):
                    cell_text = str(row[i]).strip()
                    # Case 1: cell is a known label, next cell is the value
                    if i + 1 < len(row):
                        target_field = _TABLE_LABEL_MAP.get(cell_text)
                        if target_field:
                            cell_value = str(row[i + 1]).strip()
                            if cell_value and cell_value != "-":
                                if target_field not in table_fields:
                                    table_fields[target_field] = cell_value
                            i += 2
                            continue
                    # Case 2: cell starts with a known label (merged label+value)
                    matched_prefix = False
                    for prefix in _TABLE_LABEL_PREFIXES:
                        if cell_text.startswith(prefix) and len(cell_text) > len(prefix):
                            target_field = _TABLE_LABEL_MAP[prefix]
                            cell_value = cell_text[len(prefix):].strip()
                            if cell_value and target_field not in table_fields:
                                table_fields[target_field] = cell_value
                            matched_prefix = True
                            break
                    i += 1
    return table_fields


def extract_fields(filename: str, full_text: str, result_json, page_count: int, *, file_path: str = "") -> dict:
    """
    核心业务：从 OCR 结果中提取关键业务字段（用于后续的报表生成）。
    
    采用"探针式"和"计分卡"的多维度规则引擎，由于不同类型的机构发文在排版（如文号在左上还是居中、日期靠底部还是页眉）
    和字体特征上有很大差异，这里使用了 `score` 加权机制来评判最可能的字段归属。

    Args:
        filename: 供兜底使用的文件名（可能包含项目编号或特征前缀）。
        full_text: 未经排版梳理的全部提取文本流。
        result_json: 由 ocr_engine 输出的带版面、行列和置信度坐标的富文本对象。
        page_count: 总页数（用于确定首页和尾页位置的边界条件）。
        file_path: 文件完整路径，用于从文件夹层级解析档号。
    """
    fields = {h: "" for h in HEADERS}
    fields["页数"] = str(page_count) if page_count else ""

    if not full_text:
        full_text = ""

    lines = []
    for raw in full_text.split('\n'):
        clean = _clean_line_text(raw)
        if not clean or re.match(r'^---\s*第\s*\d+\s*页\s*---$', clean):
            continue
        lines.append(clean)

    # --- 档号：从文件名 + 文件夹路径提取 ---
    fields["档号"] = _extract_archive_number(filename, file_path)

    # --- 文号 ---
    items = _collect_items(result_json, full_text)
    fields["文号"] = _extract_doc_no(items, lines)

    # --- 题名 ---
    fields["题名"] = _extract_title(items, lines)

    # --- 责任者 ---
    fields["责任者"] = _extract_responsible(items, fields["文号"])

    # --- 日期 ---
    fields["日期"] = _extract_date(items, lines)

    # --- 密级 ---
    fields["密级"] = _extract_classification(items, full_text)

    # --- 表格数据（归档页/公文处理单的结构化字段） ---
    # 策略：如果表格提取到 3+ 个字段 → 该页是归档表单，表格值可信度高，可覆盖噪声
    #       否则仅作为 fallback 补充空字段（源文件内容优先）
    table_fields = _extract_from_table_data(result_json)
    is_archive_form = len(table_fields) >= 3
    for field_name, value in table_fields.items():
        if not value or field_name not in fields:
            continue
        if is_archive_form:
            # 归档表单：表格数据覆盖（但保留档号等非表格字段）
            if field_name != "档号":
                fields[field_name] = value
        else:
            # 非归档表单：仅补充空字段
            if not fields[field_name]:
                fields[field_name] = value

    return fields

def resolve_excel_output_path(output_path: str) -> str:
    raw = (output_path or '').strip()
    p = Path(raw)
    if raw.endswith(('\\', '/')) or (p.exists() and p.is_dir()) or not p.suffix:
        p = p / DEFAULT_EXCEL_NAME
    elif p.suffix.lower() == '.xls':
        p = p.with_suffix('.xlsx')
    p.parent.mkdir(parents=True, exist_ok=True)
    return str(p)

def init_excel(output_path: str) -> str:
    """创建 Excel 文件，写入表头，返回路径"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "归档文件目录"

    ws.merge_cells('A1:H1')
    ws['A1'] = '归档文件目录'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    col_widths = {'A': 20, 'B': 25, 'C': 15, 'D': 50, 'E': 12, 'F': 6, 'G': 8, 'H': 15}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)
    logger.info("创建归档目录 Excel: %s", output_path)
    return output_path

def clear_excel_data(output_path: str):
    """清空 Excel 数据行（保留标题行和表头行），用于每次批量写入前重置"""
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    # 删除第3行及以后所有数据行
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    wb.save(output_path)
    logger.info("已清空归档目录数据行: %s", output_path)

def append_to_excel(output_path: str, fields: dict):
    """向 Excel 追加一行数据"""
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    next_row = ws.max_row + 1
    for col_idx, header in enumerate(HEADERS, 1):
        ws.cell(row=next_row, column=col_idx, value=fields.get(header, ""))
    wb.save(output_path)
