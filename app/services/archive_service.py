import logging
from pathlib import Path

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import ArchiveRecord
from app.services.excel_export import append_to_excel, init_excel


logger = logging.getLogger(__name__)


async def save_archive_record(
    db: AsyncSession,
    task_id: int | None,
    batch_id: str,
    batch_folder: str,
    fields: dict,
) -> ArchiveRecord:
    record: ArchiveRecord | None = None
    if task_id is not None:
        result = await db.execute(select(ArchiveRecord).where(ArchiveRecord.task_id == task_id))
        record = result.scalar_one_or_none()

    if record is None:
        record = ArchiveRecord(task_id=task_id)
        db.add(record)

    record.batch_id = batch_id
    record.batch_folder = batch_folder
    record.archive_no = fields.get("档号", "") or ""
    record.doc_no = fields.get("文号", "") or ""
    record.responsible = fields.get("责任者", "") or ""
    record.title = fields.get("题名", "") or ""
    record.date = fields.get("日期", "") or ""
    record.pages = fields.get("页数", "") or ""
    record.classification = fields.get("密级", "") or ""
    record.remarks = fields.get("备注", "") or ""

    await db.commit()
    await db.refresh(record)
    return record


async def get_archive_records(
    db: AsyncSession,
    folder: str = "",
    batch_id: str = "",
    page: int = 1,
    page_size: int = 200,
):
    query = select(ArchiveRecord)
    count_query = select(func.count()).select_from(ArchiveRecord)

    if folder:
        query = query.where(ArchiveRecord.batch_folder == folder)
        count_query = count_query.where(ArchiveRecord.batch_folder == folder)
    if batch_id:
        query = query.where(ArchiveRecord.batch_id == batch_id)
        count_query = count_query.where(ArchiveRecord.batch_id == batch_id)

    total = (await db.execute(count_query)).scalar() or 0
    query = query.order_by(ArchiveRecord.created_at.asc()).offset((page - 1) * page_size).limit(page_size)
    records = (await db.execute(query)).scalars().all()
    return records, total


def records_to_excel(records: list[ArchiveRecord], output_path: str) -> str:
    init_excel(output_path)
    for record in records:
        append_to_excel(
            output_path,
            {
                "档号": record.archive_no or "",
                "文号": record.doc_no or "",
                "责任者": record.responsible or "",
                "题名": record.title or "",
                "日期": record.date or "",
                "页数": record.pages or "",
                "密级": record.classification or "",
                "备注": record.remarks or "",
            },
        )
    logger.info("Exported %s archive record(s) to %s.", len(records), output_path)
    return output_path


async def import_from_excel(db: AsyncSession, file_path: str, batch_id: str = "") -> int:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File does not exist: {file_path}")

    extension = path.suffix.lower()
    rows: list[dict] = []

    if extension == ".xlsx":
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        header_row = None
        headers: list[str] = []
        for row_index in range(1, min(6, worksheet.max_row + 1)):
            values = [str(worksheet.cell(row_index, column).value or "").strip() for column in range(1, 10)]
            if "档号" in values or "文号" in values:
                header_row = row_index
                headers = values
                break
        if header_row is None:
            raise ValueError("Could not find the header row in the workbook.")
        for row_index in range(header_row + 1, worksheet.max_row + 1):
            row = {
                headers[column - 1]: str(worksheet.cell(row_index, column).value or "").strip()
                for column in range(1, min(len(headers) + 1, 10))
            }
            if any(value for value in row.values()):
                rows.append(row)
    elif extension == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError("Reading .xls files requires xlrd==1.2.0.") from exc

        workbook = xlrd.open_workbook(file_path)
        worksheet = workbook.sheet_by_index(0)
        header_row = None
        headers = []
        for row_index in range(min(6, worksheet.nrows)):
            values = [str(worksheet.cell_value(row_index, column)).strip() for column in range(min(9, worksheet.ncols))]
            if "档号" in values or "文号" in values:
                header_row = row_index
                headers = values
                break
        if header_row is None:
            raise ValueError("Could not find the header row in the workbook.")
        for row_index in range(header_row + 1, worksheet.nrows):
            row = {
                headers[column]: str(worksheet.cell_value(row_index, column)).strip()
                for column in range(min(len(headers), worksheet.ncols))
            }
            if any(value for value in row.values()):
                rows.append(row)
    else:
        raise ValueError(f"Unsupported file extension: {extension}")

    folder = str(path.parent)
    count = 0
    for row in rows:
        db.add(
            ArchiveRecord(
                task_id=None,
                batch_id=batch_id or f"import_{path.stem}",
                batch_folder=folder,
                archive_no=row.get("档号", ""),
                doc_no=row.get("文号", ""),
                responsible=row.get("责任者", ""),
                title=row.get("题名", ""),
                date=row.get("日期", ""),
                pages=row.get("页数", ""),
                classification=row.get("密级", ""),
                remarks=row.get("备注", ""),
            )
        )
        count += 1

    await db.commit()
    logger.info("Imported %s archive record(s) from %s.", count, file_path)
    return count
